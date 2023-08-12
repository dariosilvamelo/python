"""
https://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
pipenv install openpyxl
"""
from random import uniform

import openpyxl

# Exemplo 01:

pedidos = openpyxl.load_workbook('.\\323-Openpyxl-Excel\\pedidos.xlsx')
nome_planilhas = pedidos.sheetnames
# Retorna um a lista das planilhas contidas no arquivo(abas)
# print(nome_planilhas)

planilha1 = pedidos['Página1']

for linha in range(5, 16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

pedidos.save('.\\323-Openpyxl-Excel\\nova_planilha_01.xlsx')


# Exemplo 02:
planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

for linha in range(1, 11):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

for linha in range(1, 11):
    planilha2.cell(
        linha, 1).value = f'Luiz {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(
        linha, 2).value = f'Otávio {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(
        linha, 3).value = f'Joãozinho {linha} {round(uniform(10, 100), 2)}'

planilha.save('.\\323-Openpyxl-Excel\\nova_planilha_02.xlsx')
